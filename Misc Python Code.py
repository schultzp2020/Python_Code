try:
	import pandas as pd
	import sys
	import os
	import datetime
	from win32com.client import Dispatch
	from os import listdir
	from os.path import isfile, join
	from openpyxl.utils import get_column_letter
	from openpyxl.styles.borders import Border, Side
	from openpyxl import load_workbook, styles
	from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception as e:
	print(e)
	

def get_most_recent_file(pwd):
    current_date = datetime.datetime.now()
    current_date = datetime.date(current_date.year,
                                 current_date.month, current_date.day)

    pwd_files = [f for f in listdir(pwd) if isfile(join(pwd, f))]

    if not len(pwd_files):
        x = datetime.datetime.now()
        pwd = f"{os.getcwd()}\\{str(x.year - 1)}\\"
        pwd_files = [f for f in listdir(pwd) if isfile(join(pwd, f))]

    most_recent_file = ""
    most_recent_file_date = 1000000

    for file_ in pwd_files:
        try:
            file_date = file_.split(" ")[3]
            file_date = file_date.split(".")[0]

            password_reset_date = datetime.date(
                int(file_date.split("-")[0]), int(file_date.split("-")[1]), int(file_date.split("-")[2]))
            days_passed = (current_date - password_reset_date).days
            if(days_passed < most_recent_file_date):
                most_recent_file_date = days_passed
                most_recent_file = file_
        except Exception:
            pass

    return pwd + most_recent_file


# Create the excel workbook
def create_excel_workbook(dfs, ws_names, filepath, old_filepath):
    if(filepath == old_filepath):
        print("Already created this month's report!")
        input('Press "Enter" to close...')
        sys.exit()

    # Create a worksheet for each dataframe in dfs
    try:
        writer = pd.ExcelWriter(filepath, engine="xlsxwriter")
        for i, df in enumerate(dfs):
            df.to_excel(writer, sheet_name=ws_names[i], index=False)
        writer.save()
    except Exception as e:
        print(e)
        input('Press "Enter" to close...')

    wb = load_workbook(filepath)

    # Create table for each worksheet in the workbook
    for table_number, ws_name in enumerate(ws_names):
        ws = wb[ws_name]
        ws = create_table(ws, table_number)

    # Save the workbook
    wb.save(filepath)

  
def create_table(ws, table_number):
    # Create a table in the worksheet
    tab = Table(displayName=f"Table{table_number}", ref="A1:" +
                get_column_letter(ws.max_column) + str(ws.max_row))
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

    # Create border object
    thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    # Extend the width of the columns
    column_widths = []
    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            try:
                column_widths[i] = max(column_widths[i], len(str(cell.value)))
                cell.border = thin_border
            except IndexError:
                column_widths.append(len(str(cell.value)))

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width

    return ws


def create_chart(df, filepath, old_filepath):
    xl = Dispatch("Excel.Application")

    wb1 = xl.Workbooks.Open(Filename=old_filepath)
    wb2 = xl.Workbooks.Open(Filename=filepath)

    try:
        ws1 = wb1.Worksheets(1)
        ws1.Copy(Before=wb2.Worksheets(1))
    except:
        pass

    wb1.Close(SaveChanges=False)
    wb2.Close(SaveChanges=True)
    xl.Quit()

    # Change data in Averages worksheet
    wb = load_workbook(filepath)
    ws1 = wb["All Results"]
    ws2 = wb["Averages"]

    for i in range(36, 49):
        ws2[f"A{i}"] = ws2[f"A{i + 1}"].value
        ws2[f"B{i}"] = ws2[f"B{i + 1}"].value

    x = datetime.datetime.now()
    date = f"{x.month}/{x.day}/{x.year}"
    ws2["A49"] = date
    ws2["A49"].alignment = styles.Alignment(horizontal='right')
    ws2["B49"] = df.shape[0]
    wb.save(filepath)